import streamlit as st
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tempfile
import os

# =======================
# STREAMLIT UI SETUP
# =======================
st.set_page_config(page_title="PIN Code Generator", layout="wide")

st.title("🔧 PIN Code Generator")
st.markdown("Upload an Excel file to generate PIN codes based on the configured mappings.")

# File uploader
uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx'])

if uploaded_file is not None:
    # Create a temporary file to save the uploaded content
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        tmp_file.write(uploaded_file.getvalue())
        file_path = tmp_file.name
    
    # Read the Excel file
    df = pd.read_excel(file_path)
    df.columns = df.columns.str.strip()
    
    # Create columns for layout
    col1, col2 = st.columns([1, 3])
    
    with col1:
        st.info(f"📊 File: {uploaded_file.name}")
        st.info(f"📈 Rows: {len(df)}")
        
        if st.button("Generate PIN Codes", type="primary"):
            with st.spinner("Processing..."):
                # =======================
                # HELPER FUNCTIONS
                # =======================
                def contains_map(text, mapping, default=""):
                    if pd.isna(text):
                        return default
                    text = str(text)
                    for key, value in mapping.items():
                        if key in text:
                            return value
                    return default

                def extract_after_dash(text, index):
                    if pd.isna(text) or "-" not in str(text):
                        return ""
                    part = str(text).split("-", 1)[1]
                    return part[index] if len(part) > index else ""

                # =======================# ===== DO NOT MODIFY START =====
                #1 MODEL NUMBER - code
                # =======================
                model_map = {
                    "-5": "05", "-10": "10", "-18": "18", "-21": "21",
                    "-33": "33", "-35": "35", "-41": "41",
                    "-77": "77", "-78": "78", "-80": "80", "-28": "28",
                }
                df["Model Number-Code"] = df["Model Number"].apply(lambda x: contains_map(x, model_map))

                # =======================
                #2 SIZE - code
                # =======================
                size_map = {
                    "0.5 x": "05", "0.7 x": "75", "1 x": "01", "1.5 x": "15",
                    "2 x": "02", "3 x": "03", "4 x": "04", "6 x": "06",
                    "8 x": "08", "10 x": "10", "12 x": "12", "14 x": "14",
                    "16 x": "16", "18 x": "18", "20 x": "20", "24 x": "24",
                    "26 x": "26", "28 x": "28", "30 x": "30",
                    "36 x": "36", "40 x": "40", "42 x": "42", "48 x": "48"
                }
                df["In x Body x Out Size-Code"] = df["In x Body x Out Size"].apply(lambda x: contains_map(x, size_map))

                # =======================
                #3 RATING CLASS  - code
                # =======================
                rating_map = {
                    "150": "1", "300": "2", "600": "3",
                    "900": "4", "1500": "5", "2500": "6"
                }
                df["Rating Class-Code"] = df["Rating Class"].apply(lambda x: contains_map(x, rating_map))

                # =======================
                #4 END CONNECTION - code
                # =======================
                end_conn_map = {
                    "RF": "RF", "FF": "FF", "RTJ": "RJ",
                    "Lugged": "LG", "BW": "BW", "SW": "SW"
                }
                df["End Connection-Code"] = df["End Connection"].apply(lambda x: contains_map(x, end_conn_map))

                # =======================
                #5 BODY MATERIAL - code
                # =======================
                body_mat_map = {
                    "WCC": "A", "LCC": "B", "A105": "C", "LF2": "D",
                    "CF8 ": "E", "CF3 ": "F", "CF8M": "G", "CF3M": "H",
                    "Duplex": "I", "Super Duplex": "J", "Aluminum Bronze": "K" , "12MW" : "L" , "C95800" : "M"
                }
                df["Body Material-Code"] = df["Body Material"].apply(lambda x: contains_map(x, body_mat_map))

                # =======================
                #6 BODY STUDS    - code
                # =======================
                df["Body Studs-Code"] = df["Body Studs"].apply(
                    lambda x: "2" if pd.notna(x) and "coat" in str(x).lower() else "1"
                )

                # =======================
                #7 BONNET TYPE   - code
                # =======================
                bonnet_map = {
                    "Standard": "ST",
                    "Extended": "EB",
                    "Finned": "FB"
                }
                df["Bonnet Type-Code"] = df["Bonnet Type"].apply(lambda x: contains_map(x, bonnet_map, "NA"))

                # =======================
                #8 ACTUATOR MODEL - code
                # =======================
                act_model_map = {
                    "Top Mounted Handwheel": "20",
                    "87": "87", "88": "88",
                    "51": "51", "52": "52", "53": "53",
                    "37": "37", "38": "38",
                    "Electrical Linear": "EL",
                    "Electrical Rotary": "ER"
                }
                df["Actuator Model-Code"] = df["Actuator Model"].apply(lambda x: contains_map(x, act_model_map))

                # =======================
                #9 ACTUATOR SIZE - code
                # =======================
                act_size_map = {
                    "6": "A", "12": "B", "16": "C", "20": "D",
                    "23L": "F", "23": "E", "11": "G", "13": "H",
                    "15": "I", "18": "J", "24": "K",
                    "Electric": "L", "10": "M"
                }
                df["Actuator Size-Code"] = df["Actuator Size"].apply(lambda x: contains_map(x, act_size_map))

                # =======================
                #10 PLUG MATERIAL - code
                # =======================
                def plug_material_code(text):
                    if pd.isna(text):
                        return ""
                    t = str(text)
                    if "316" in t and ("Hard" in t or "HF" in t):
                        return "3"
                    if "316" in t:
                        return "2"
                    if "410" in t:
                        return "1"
                    if "CA6NM" in t and ("Plating" in t or "coat" in t):
                        return "5"
                    if "CA6NM" in t:
                        return "4"
                    if "31254" in t:
                        return "6"
                    if "C276" in t:
                        return "7"
                    if "Monel" in t:
                        return "8"  
                    if "Stellite" in t:
                        return "9"
                    return ""

                df["Plug Material-Code"] = df["Plug Material"].apply(plug_material_code)

                # =======================
                #11 Plug Type-Code
                # =======================
                df["Plug Type-Code"] = df["Model Number"].apply(lambda x: extract_after_dash(x, 2))

                # =======================
                #12 Trim Type-Code
                # =======================
                df["Trim Type-Code"] = df["Model Number"].apply(lambda x: extract_after_dash(x, 3))

                # =======================
                #13 Seat Type-Code
                # =======================
                df["Seat Type-Code"] = df["Model Number"].apply(lambda x: extract_after_dash(x, 4))

                # =======================
                #13 TRIM CHARACTERISTIC - code
                # =======================
                trim_char_map = {
                    "Contoured"	: "A",
                    "Linear" : "B",
                    "Equal Percent" : "C",
                    "Modified Percentage" :"D",
                    "Quick Opening"	:	"E",
                    "Anti-Cavitation 1 Stage - Linear"	: "F",
                    "Anti-Cavitation 1 Stage - Equal Percentage"	: "G",
                    "Anti-Cavitation 2 Stage - Linear"	:"H",
                    "Anti-Cavitation 2 Stage - Equal Percentage"	:"I",
                    "50 % LODB 1 Stage Equal % + 50% Contoured"	:"J",
                    "LoDB 1 Stage - Linear"	:"K",
                    "LoDB 2 Stage - Linear"	:"L",
                    "LoDB 1 Stage - Equal Percentage"	:"M",
                    "LoDB 2 Stage - Equal Percentage"	:"N",
                    "Antisurge Lo dB 1 Stage"	:"O",
                    "Antisurge Lo dB 2 Stage"	:"P",
                    "LoDB 1 Stage - Close clearance Linear"	:"Q",
                    "LoDB 2 Stage - Close clearance Linear"	:"R"   ,
                }
                df["Trim Characteristic-Code"] = df["Trim Characteristic"].apply(lambda x: contains_map(x, trim_char_map))

                ## Helper function for mapping lookup
                def get_mapping(model, pos, mapping_dict):
                    if pd.isna(model):
                        return ""
                    m = str(model)
                    key = extract_after_dash(m, pos)
                    for prefix, mapping in mapping_dict.items():
                        if prefix in m:
                            return mapping.get(key, "")
                    return ""

                # =======================
                # 14 PLUG TYPE DESCRIPTION
                # =======================
                plug_mappings = {
                    "-41": {"0": "Undefined", "3": "Pressure energized PTFE seal ring", "4": "With pilot",
                            "5": "Metal seal ring", "6": "PTFE seal ring", "7": "HT metal seal ring", "9": "Graphite seal ring"},

                    "-21": {"0": "Undefined", "1": "Contoured", "3": "Close Clearance Lo-dB/Anti-cavitation",
                            "5": "Over Travel", "6": "Soft Seat", "7": "Single Stage Lo-dB/Anti-cavitation",
                            "8": "Double Stage Anti-cavitation", "9": "Double Stage Lo-dB"},

                    "-10": {"0": "Undefined", "1": "Double Seat"},

                    "-18": {"0": "Undefined", "1": "Axial Flow High Resistance (Downseating)"},

                    "-78": {"0": "Undefined", "1": "Axial Flow High Resistance (Downseating)"},

                    "-80": {"0": "Undefined", "3": "Top and Port Guided"},

                    "-33": {"0": "Triple Offset"},  # uses y

                    "-35": {"0": "Self-aligning eccentrically rotatingt"},

                    "-28": {"0": "3.8, Linear", "1": "2.3, Linear", "2": "1.2, Linear", "3": "0.6, Linear",
                            "4": "0.25, Linear", "5": "0.1, Linear", "6": "0.05, Modified Linear", "7": "0.025, Modified Linear",
                            "8": "0.01, Modified Linear", "9": "0.004, Modified Linear"},  # uses y

                    "-77": {"0": "Undefined", "1": "Trim A: 9-stage unbalanced", "2": "Trim B: 5-stage unbalanced",
                            "3": "Trim C: 1-stage unbalanced", "4": "Trim X: 5-stage partially balanced", "5": "Trim Y: 3-stage unbalanced"},
                }

                def plug_type_desc(model):
                    if pd.isna(model):
                        return ""
                    m = str(model)
                    # Decide if we need x or y
                    x, y = extract_after_dash(m, 2), extract_after_dash(m, 3)
                    if "-33" in m or "-28" in m:
                        return get_mapping(m, 3, plug_mappings)
                    return get_mapping(m, 2, plug_mappings)

                df["Plug Type-Des"] = df["Model Number"].apply(plug_type_desc)


                # =======================
                # 15 TRIM TYPE DESCRIPTION
                # =======================
                trim_mappings = {
                    "-41": {"0": "Undefined", "1": "Standard cage / Linear", "2": "Standard cage / Equal percentage",
                            "3": "Lo-dB® / Anticavitation single stage / Linear", "4": "Lo-dB® single stage with diffuser / Linear",
                            "5": "Lo-dB® double stage / Linear", "6": "VRT (stack) Type S / Linear", "7": "VRT (stack partial) Type S / modified percentage",
                            "8": "VRT (cage) Type C / Linear", "9": "Anticavitation double stage / Linear (1)", "A": "High Capacity Linear",
                            "B": "High Capacity Equal %", "C": "High Capacity Lo-DB / Anti-Cav"},

                    "-21": {"4": "Quick Change", "5": "Threaded"},

                    "-18": {str(i): f"Trim {chr(65 + (i-1)//3)}, {'Balanced' if i<=6 else 'Unbalanced'} {'Hard' if i%3==1 else 'Soft' if i%3==2 else 'Hard Seat' if i>6 else 'Soft Seat'}" for i in range(10)},
                    "-78": {str(i): f"Trim {chr(65 + (i-1)//3)}, {'Balanced' if i<=6 else 'Unbalanced'} {'Hard' if i%3==1 else 'Soft' if i%3==2 else 'Hard Seat' if i>6 else 'Soft Seat'}" for i in range(10)},

                    "-77": {"0": "Optional Trim", "1": "Bottom entry; outlet spool", "2": "Top entry; bolted bonnet", "3": "Compact Top entry; bolted bonnet"},

                    "-80": {"0": "Combined design", "1": "Diverting design"},

                    "-28": {"0": "Metal seat"},

                    "-33": {"0": "Metal + Graphite Laminated"},

                    "-35": {"1": "Metal Seat", "2": "Soft Seat", "3": "Metal Seat w/ Differential Velocity Trim", "4": "Soft Seat w/ Differential Velocity Trim"},
                }

                def trim_type_desc(model):
                    if pd.isna(model):
                        return ""
                    m = str(model)
                    x, y = extract_after_dash(m, 3), extract_after_dash(m, 4)
                    if "-41" in m:
                        return trim_mappings["-41"].get(x, "")
                    return get_mapping(m, 4, trim_mappings)

                df["Trim Type-Des"] = df["Model Number"].apply(trim_type_desc)

                # =======================
                # 16 PIN CODE
                # =======================
                pin_columns = [
                    "Model Number-Code",
                    "In x Body x Out Size-Code",
                    "Rating Class-Code",
                    "End Connection-Code",
                    "Body Material-Code",
                    "Body Studs-Code",
                    "Bonnet Type-Code",
                    "Actuator Model-Code",
                    "Actuator Size-Code",
                    "Plug Material-Code",
                    "Trim Type-Code",
                    "Seat Type-Code",
                    "Trim Characteristic-Code"
                ]

                df["PIN-Code"] = df[pin_columns].fillna("").astype(str).agg("".join, axis=1)
                df["PIN-Code-Length"] = df["PIN-Code"].astype(str).str.len()

                # =======================
                # 17 PIN DESCRIPTION
                # =======================
                desc_columns = [
                    "Model Number",
                    "In x Body x Out Size",
                    "Rating Class",
                    "End Connection",
                    "Body Material",
                    "Body Studs",
                    "Bonnet Type",
                    "Actuator Model",
                    "Actuator Size",
                    "Plug Material",
                    "Trim Type-Des",
                    "Plug Type-Des",
                    "Seat Type",
                    "Trim Characteristic"
                ]

                df["PIN-Code description"] = df[desc_columns].fillna("").astype(str).agg(", ".join, axis=1)

                # =======================# ===== DO NOT MODIFY END =====
                
                # Create output file in temp directory
                output_file = tempfile.NamedTemporaryFile(delete=False, suffix='_PIN_Generated.xlsx').name
                df.to_excel(output_file, index=False)

                # =======================
                # FORMATTING
                # =======================
                wb = load_workbook(output_file)
                ws = wb.active

                light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

                # =======================
                # 18 PIN Length
                # =======================
                header_map = {
                    ws.cell(row=1, column=col).value: col
                    for col in range(1, ws.max_column + 1)
                }

                if "PIN-Code-Length" in header_map:
                    col_idx = header_map["PIN-Code-Length"]

                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col_idx)
                        try:
                            if int(cell.value) < 18:
                                cell.fill = light_blue_fill
                        except:
                            pass

                green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                new_columns = [
                    "Model Number-Code",
                    "In x Body x Out Size-Code",
                    "Rating Class-Code",
                    "End Connection-Code",
                    "Body Material-Code",
                    "Body Studs-Code",
                    "Bonnet Type-Code",
                    "Actuator Model-Code",
                    "Actuator Size-Code",
                    "Plug Material-Code",
                    "Plug Type-Code",
                    "Trim Type-Code",
                    "Seat Type-Code",
                    "Trim Characteristic-Code",
                    "Trim Type-Des",
                    "Plug Type-Des",
                    "PIN-Code",
                    "PIN-Code description"
                ]

                header_map = {
                    ws.cell(row=1, column=col).value: col
                    for col in range(1, ws.max_column + 1)
                }

                for col_name in new_columns:
                    if col_name not in header_map:
                        continue

                    col_idx = header_map[col_name]
                    ws.cell(row=1, column=col_idx).fill = green_fill

                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col_idx)
                        if cell.value is None or str(cell.value).strip() == "":
                            cell.fill = yellow_fill

                wb.save(output_file)
                
                # Store the output file path in session state for download
                st.session_state['output_file'] = output_file
                st.session_state['df_result'] = df
                st.session_state['processed'] = True
    
    with col2:
        if st.session_state.get('processed', False):
            st.success("✅ Processing completed!")
            
            # Display results in expandable sections
            with st.expander("📋 View Generated Data", expanded=True):
                st.dataframe(st.session_state['df_result'], use_container_width=True)
            
            # Statistics
            col_stats1, col_stats2, col_stats3 = st.columns(3)
            with col_stats1:
                st.metric("Total Rows", len(st.session_state['df_result']))
            with col_stats2:
                valid_pins = st.session_state['df_result']['PIN-Code-Length'].ge(18).sum()
                st.metric("Valid PINs (≥18 chars)", valid_pins)
            with col_stats3:
                invalid_pins = st.session_state['df_result']['PIN-Code-Length'].lt(18).sum()
                st.metric("Invalid PINs (<18 chars)", invalid_pins)
            
            # Download button
            with open(st.session_state['output_file'], 'rb') as f:
                st.download_button(
                    label="📥 Download Generated Excel File",
                    data=f,
                    file_name=f"PIN_Generated_{uploaded_file.name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            # Clean up temp files
            try:
                os.unlink(file_path)
                os.unlink(st.session_state['output_file'])
            except:
                pass

else:
    st.info("👆 Please upload an Excel file to begin.")

# Footer
st.markdown("---")
st.markdown("🔧 PIN Code Generator - v1.0")